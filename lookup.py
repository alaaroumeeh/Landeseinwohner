import openpyxl,tabulate
from sortedcontainers import SortedList

#function definitions of user commands

def show(arg):
    #arg : Land name OR Cell, if arg != ""
    if arg == "":
        print(tabulate.tabulate(ws_list,tablefmt="grid"))
        
    elif arg in lands: #arg is land name
        landidx = lands.index(arg)
        partial_list = [headers[:],ws_list[landidx+1][:]]
        
        print(tabulate.tabulate(partial_list,tablefmt="grid"))

    else: #arg is cell
        try:
            print(ws[arg].value)    
        except Exception:
            print(f"Invalid land name or cell index : '{arg}'")


def add(args):
    #args[i] corresponds to a value for headers[i]
    lands.add(args[0])
    landidx = lands.index(args[0])
    ws_list.insert(landidx+1,args)
    ws.insert_rows(landidx+2)
    for i in range(len(args)): #len(args) = 6 always
        ws.cell(row=landidx+2,column=i+1,value=args[i])
    print(f"Row added: {args}")

def remove(args):
    for arg in args:
        if arg in lands:
            landidx = lands.index(arg)
            lands.remove(arg)
            ws.delete_rows(landidx+2)
            removed_row = ws_list.pop(landidx+1)
            print(f"Removed row: {removed_row}") #works when all joined cells have strings
        else:
            print(f"Unable to remove row: No such land name '{arg}'")

def modify(arg):
    #arg is land name
        landidx = lands.index(arg)
        for i in range(1,6):
            new_val = input(f"{headers[i]}: ").strip()
            if new_val != "":
                ws_list[landidx+1][i] = new_val
                ws.cell(row=landidx+2,column=i+1,value=new_val)
        print(f"Row modified: {[cell.value for cell in ws[landidx+2]]}")

#load workbook and active worksheet

wb = openpyxl.load_workbook("example.xlsx")
ws = wb.active


#read land names in worksheet into a sorted list

lands = SortedList(ws.cell(row=i,column=1).value for i in range(2,ws.max_row+1))

#read whole worksheet values into 2D list
ws_list = []

for row in ws.values:
        ws_list.append(list(row))

#store headers in list (same as ws_list[0])

headers = ["Landesname","Nationalität","Einwohner (m)","Einwohnerin (f)", "Einwohner (m pl.)","Einwohnerinnen (f pl.)"]
fields = ["Command","Landesname","Nationalität","Einwohner (m)","Einwohnerin (f)", "Einwohner (m pl.)","Einwohnerinnen (f pl.)"]

print(tabulate.tabulate(ws_list,tablefmt="grid"))

#prompt user for actions
#commands: show,add,remove,modify
#input format: (command) [arg1], [arg2], [arg3], ..., [arg6]

while True:
    try:
        cmd = input("/>").strip().lower()
        if cmd == "show":
            arg = input("Land name or cell index: ").strip()
            show(arg)

        elif cmd == "add":
            args = []
            correct = True
            for i in range(6):
                arg = input(f"{headers[i]}: ").strip()
                if i == 0 and arg == "":
                    print("Unable to add row: No land name provided.")
                    correct = False
                    break
                elif i == 0 and arg in lands:
                    print(f"Unable to add row: '{arg}' already exists.")
                    correct = False
                    break
                if arg == "":
                    args.append("-")
                else:
                    args.append(arg)
            if correct:
                add(args)

        elif cmd == "remove":
            args = []
            i = 1
            while True:
                arg = input(f"Land name #{i}: ").strip()
                if arg == "":
                    break
                args.append(arg)
                i += 1
            if args != []:
                remove(args)
            else:
                print("Unable to delete row: No land name provided.")

        elif cmd == "modify":
            arg = input("Select land name to modify : ").strip()
            if arg in lands:
                modify(arg)
            else:
                print("Unable to modify row : Invalid land name.")

        elif cmd == "exit" or cmd == "quit":
            break

        else:
            print(f"Unknown command: '{cmd}'")
    except KeyboardInterrupt:
        break

wb.save("example.xlsx")